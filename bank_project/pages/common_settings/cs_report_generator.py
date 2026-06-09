"""
cs_report_generator.py
---------------------
Generic Excel report generator for Common Settings test automation.

Auto-generates after every pytest run in the Common Settings module.
Parses test logs for step-level details - no test code changes needed.

Output: pages/common_settings/reports/CommonSettings_Report_YYYYMMDD_HHMMSS.xlsx
  Sheet 1: Summary        - totals, module breakdown, KPIs
  Sheet 2: Test Results   - per-test pass/fail/duration (full-row colored)
  Sheet 3: Step Details   - auto-parsed from log messages
  Sheet 4: Error Details  - only created if failures exist
  Sheet 5: Known Issues   - bugs/findings documented during testing
"""

import os
import re
import logging
import traceback
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# -- Color Palette (matching co_report_generator style) ----------
FILL_DARK   = PatternFill("solid", fgColor="1F3864")
FILL_MED    = PatternFill("solid", fgColor="2E75B6")
FILL_LIGHT  = PatternFill("solid", fgColor="D6E4F0")
FILL_ALT    = PatternFill("solid", fgColor="F5F5F5")
FILL_WHITE  = PatternFill("solid", fgColor="FFFFFF")
FILL_GREEN  = PatternFill("solid", fgColor="C6EFCE")
FILL_RED    = PatternFill("solid", fgColor="FFC7CE")
FILL_GOLD   = PatternFill("solid", fgColor="FFEB9C")

FONT_TITLE    = Font(name="Calibri", bold=True, color="FFFFFF", size=16)
FONT_SUBTITLE = Font(name="Calibri", italic=True, color="FFFFFF", size=11)
FONT_HEADER   = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
FONT_NORMAL   = Font(name="Calibri", size=10)
FONT_BOLD     = Font(name="Calibri", bold=True, size=10)
FONT_KPI_VAL  = Font(name="Calibri", bold=True, size=18, color="1F3864")
FONT_KPI_LBL  = Font(name="Calibri", bold=True, size=10, color="FFFFFF")
FONT_PASS     = Font(name="Calibri", bold=True, size=10, color="006100")
FONT_FAIL     = Font(name="Calibri", bold=True, size=10, color="9C0006")

BORDER = Border(
    left=Side("thin", color="D9D9D9"),
    right=Side("thin", color="D9D9D9"),
    top=Side("thin", color="D9D9D9"),
    bottom=Side("thin", color="D9D9D9"),
)

A_C = Alignment(horizontal="center", vertical="center", wrap_text=True)
A_L = Alignment(horizontal="left", vertical="center", wrap_text=True)


# -- Module detection map (add new Common Settings modules here) --
MODULE_MAP = {
    "uom": "UOM",
    "currency": "Currency",
    "tax": "Tax",
    "tax_category": "Tax Category",
    "item_category": "Item Category",
    "uom_conversion": "UOM Conversion",
    "hsn": "HSN",
    "payment_terms": "Payment Terms",
    "shipping": "Shipping",
}


def _detect_module(nodeid):
    """Detect module name from pytest node ID."""
    filename = nodeid.split("::")[0] if "::" in nodeid else nodeid
    filename = os.path.basename(filename).lower()
    name_part = filename.replace("test_", "").replace(".py", "")
    name_part = name_part.replace("_full_flow", "").replace("_create", "")
    name_part = name_part.replace("_edit", "").replace("_validation", "")
    for key, module_name in MODULE_MAP.items():
        if key in name_part:
            return module_name
    clean = name_part.replace("_", " ").title()
    return clean if clean else "Unknown"


# -- Sanitize (fixes Python 3.14 + openpyxl control char issue) --

def _sanitize(value):
    """Strip non-printable characters that break openpyxl on Python 3.14."""
    if not isinstance(value, str):
        if value is None:
            return ""
        return value
    return "".join(c for c in value if c >= " " or c in "\n\r\t")


# -- Style helpers (matching co_report_generator pattern) --------

def _sc(cell, font=None, fill=None, align=None, border=None):
    """Apply styles and sanitize cell value."""
    if font:   cell.font = font
    if fill:   cell.fill = fill
    if align:  cell.alignment = align
    if border: cell.border = border
    if cell.value is not None and isinstance(cell.value, str):
        cell.value = _sanitize(cell.value)


def _title_banner(ws, title, subtitle, cols=10):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=cols)
    c = ws.cell(row=1, column=1, value=title)
    _sc(c, FONT_TITLE, FILL_DARK, A_C)
    ws.row_dimensions[1].height = 36
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=cols)
    c = ws.cell(row=2, column=1, value=subtitle)
    _sc(c, FONT_SUBTITLE, FILL_MED, A_C)
    ws.row_dimensions[2].height = 22
    ws.row_dimensions[3].height = 8


def _headers(ws, row, heads, start=1):
    for i, h in enumerate(heads, start):
        c = ws.cell(row=row, column=i, value=h)
        _sc(c, FONT_HEADER, FILL_MED, A_C, BORDER)
    ws.row_dimensions[row].height = 24
    return row + 1


def _row(ws, row, vals, start=1, alt=False):
    fill = FILL_ALT if alt else FILL_WHITE
    for i, v in enumerate(vals, start):
        v = _sanitize(v) if isinstance(v, str) else v
        c = ws.cell(row=row, column=i, value=v)
        _sc(c, FONT_NORMAL, fill, A_L, BORDER)
    return row + 1


def _row_full_status(ws, row, vals, status, start=1):
    """Write a row with full-row coloring based on pass/fail status."""
    if status == "PASSED":
        fill = FILL_GREEN
        font = FONT_PASS
    elif status == "FAILED":
        fill = FILL_RED
        font = FONT_FAIL
    else:
        fill = FILL_WHITE
        font = FONT_NORMAL
    for i, v in enumerate(vals, start):
        v = _sanitize(v) if isinstance(v, str) else v
        c = ws.cell(row=row, column=i, value=v)
        _sc(c, font, fill, A_L, BORDER)
    return row + 1


def _status_cell(ws, row, col, status):
    """Color a single status cell."""
    status = _sanitize(status)
    c = ws.cell(row=row, column=col, value=status)
    if status == "PASSED":
        _sc(c, FONT_PASS, FILL_GREEN, A_C, BORDER)
    elif status == "FAILED":
        _sc(c, FONT_FAIL, FILL_RED, A_C, BORDER)
    elif status == "SKIPPED":
        _sc(c, FONT_NORMAL, FILL_ALT, A_C, BORDER)
    else:
        _sc(c, FONT_NORMAL, FILL_WHITE, A_C, BORDER)


def _widths(ws, spec, mn=10, mx=50):
    for col, w in spec.items():
        ws.column_dimensions[get_column_letter(col)].width = min(max(w, mn), mx)


# -- Log parser (auto-extracts steps from log.info lines) -------

STEP_PASSED = re.compile(r">>>\s*STEP\s*(\w+)\s*PASSED[:\s]*(.*)", re.IGNORECASE)
STEP_FAILED = re.compile(r">>>\s*STEP\s*(\w+)\s*FAILED[:\s]*(.*)", re.IGNORECASE)


def parse_steps_from_logs(log_messages):
    """
    Parse test log messages to extract step-level details.
    Looks for patterns like:
      >>> STEP 1 PASSED: UOM created and verified
      >>> STEP 5 PASSED: History verified with updated data
    Returns list of dicts: [{step, status, detail}, ...]
    """
    steps = []
    for msg in log_messages:
        msg = msg.strip()
        m = STEP_PASSED.search(msg)
        if m:
            steps.append({
                "step": _sanitize("Step " + m.group(1)),
                "status": "PASSED",
                "detail": _sanitize(m.group(2).strip()),
            })
            continue
        m = STEP_FAILED.search(msg)
        if m:
            steps.append({
                "step": _sanitize("Step " + m.group(1)),
                "status": "FAILED",
                "detail": _sanitize(m.group(2).strip()),
            })
    return steps


# -- Main report generator --------------------------------------

def generate_cs_report(results, output_dir, issues=None):
    """
    Generate the Common Settings Excel report.

    Args:
        results: list of dicts (from CSReportStore.results)
        output_dir: directory to save the xlsx file
        issues: optional list of issue dicts (from CSReportStore.known_issues)

    Returns:
        str: absolute path to the generated file
    """
    os.makedirs(output_dir, exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    fp = os.path.join(output_dir, "CommonSettings_Report_" + ts + ".xlsx")

    wb = Workbook()
    _build_summary(wb, results)
    _build_test_results(wb, results)
    _build_step_details(wb, results)

    has_failures = any(r.get("status") == "FAILED" for r in results)
    if has_failures:
        _build_error_details(wb, results)

    if issues:
        _build_known_issues(wb, issues)

    wb.save(fp)
    return fp


# -- Sheet builders ---------------------------------------------

def _build_summary(wb, results):
    ws = wb.active
    ws.title = "Summary"
    ws.sheet_properties.tabColor = "1F3864"

    total   = len(results)
    passed  = sum(1 for r in results if r.get("status") == "PASSED")
    failed  = sum(1 for r in results if r.get("status") == "FAILED")
    skipped = sum(1 for r in results if r.get("status") == "SKIPPED")
    rate    = (passed / total * 100) if total else 0
    total_duration = sum(r.get("duration", 0) for r in results)

    _title_banner(ws, "COMMON SETTINGS - TEST AUTOMATION REPORT",
                  "Generated: " + datetime.now().strftime("%d-%b-%Y %H:%M"), cols=12)

    # KPI row
    r = 4
    kpis = [
        ("Total Tests",    str(total),             FILL_LIGHT),
        ("Passed",         str(passed),            FILL_GREEN),
        ("Failed",         str(failed),            FILL_RED),
        ("Skipped",        str(skipped),           FILL_ALT),
        ("Pass Rate",
         str(int(rate)) + "%",
         FILL_GREEN if rate == 100 else FILL_GOLD if 0 < rate < 100 else FILL_RED),
        ("Duration",       _fmt_duration(total_duration), FILL_LIGHT),
    ]
    for i, (label, val, fill) in enumerate(kpis):
        col = i * 2 + 1
        ws.merge_cells(start_row=r, start_column=col, end_row=r, end_column=col + 1)
        _sc(ws.cell(row=r, column=col, value=label), FONT_KPI_LBL, FILL_MED, A_C, BORDER)
        ws.cell(row=r, column=col + 1).border = BORDER
        ws.merge_cells(start_row=r + 1, start_column=col, end_row=r + 1, end_column=col + 1)
        _sc(ws.cell(row=r + 1, column=col, value=val), FONT_KPI_VAL, fill, A_C, BORDER)
        ws.cell(row=r + 1, column=col + 1).border = BORDER
    ws.row_dimensions[r].height = 24
    ws.row_dimensions[r + 1].height = 36

    # Module breakdown table
    r = 8
    r = _headers(ws, r, ["#", "Module", "Tests", "Passed", "Failed", "Skipped", "Duration", "Status"])

    modules = {}
    for res in results:
        mod = res.get("module", "Unknown")
        if mod not in modules:
            modules[mod] = {"total": 0, "passed": 0, "failed": 0, "skipped": 0, "duration": 0.0}
        modules[mod]["total"] += 1
        st = res.get("status", "")
        if st == "PASSED":   modules[mod]["passed"] += 1
        elif st == "FAILED": modules[mod]["failed"] += 1
        elif st == "SKIPPED": modules[mod]["skipped"] += 1
        modules[mod]["duration"] += res.get("duration", 0)

    for i, (mod, stats) in enumerate(sorted(modules.items()), 1):
        alt = i % 2 == 0
        mod_status = "PASS" if stats["failed"] == 0 else "FAIL"
        vals = [i, mod, stats["total"], stats["passed"], stats["failed"],
                stats["skipped"], _fmt_duration(stats["duration"]), mod_status]
        r = _row(ws, r, vals, alt=alt)
        _status_cell(ws, r - 1, 8, mod_status)

    _widths(ws, {1: 5, 2: 20, 3: 8, 4: 10, 5: 10, 6: 10, 7: 14, 8: 12})


def _build_test_results(wb, results):
    ws = wb.create_sheet("Test Results")
    ws.sheet_properties.tabColor = "2E75B6"

    _title_banner(ws, "TEST RESULTS", "Detailed results per test case", cols=8)

    r = 4
    r = _headers(ws, r, ["#", "Test Name", "Module", "Status", "Duration", "Timestamp", "Steps", "Error"])

    for i, res in enumerate(results, 1):
        status = res.get("status", "UNKNOWN")
        steps = res.get("steps", [])
        step_passed = sum(1 for s in steps if s.get("status") == "PASSED")
        step_str = str(step_passed) + "/" + str(len(steps)) if steps else "-"
        error_short = _truncate(res.get("error", ""), 80)

        vals = [i, res.get("test_name", ""), res.get("module", ""), status,
                _fmt_duration(res.get("duration", 0)),
                res.get("timestamp", ""), step_str, error_short]
        r = _row_full_status(ws, r, vals, status)

    _widths(ws, {1: 5, 2: 45, 3: 16, 4: 12, 5: 12, 6: 20, 7: 8, 8: 50})


def _build_step_details(wb, results):
    ws = wb.create_sheet("Step Details")
    ws.sheet_properties.tabColor = "548235"

    _title_banner(ws, "STEP DETAILS", "Step-level results (auto-parsed from test logs)", cols=5)

    r = 4
    r = _headers(ws, r, ["#", "Test Name", "Step", "Status", "Detail"])

    idx = 1
    for res in results:
        test_name = res.get("test_name", "Unknown")
        for step in res.get("steps", []):
            alt = idx % 2 == 0
            vals = [idx, test_name, step.get("step", ""), step.get("status", ""), step.get("detail", "")]
            r = _row(ws, r, vals, alt=alt)
            _status_cell(ws, r - 1, 4, step.get("status", ""))
            idx += 1

    if idx == 1:
        r = _row(ws, r, ["", "No step details captured", "", "", ""])

    _widths(ws, {1: 5, 2: 45, 3: 20, 4: 12, 5: 60})


def _build_known_issues(wb, issues):
    """Sheet 5: Known Issues / Bugs discovered during testing."""
    ws = wb.create_sheet("Known Issues")
    ws.sheet_properties.tabColor = "BF8F00"

    _title_banner(ws, "KNOWN ISSUES / BUGS",
                  str(len(issues)) + " issue(s) documented during testing", cols=9)

    r = 4
    r = _headers(ws, r, ["#", "Severity", "Module", "Category", "Description",
                          "Expected Behavior", "Actual Behavior", "Test Ref", "Status"])

    for i, issue in enumerate(issues, 1):
        alt = i % 2 == 0
        vals = [i, issue.get("severity", ""), issue.get("module", ""),
                issue.get("category", ""), issue.get("description", ""),
                issue.get("expected", ""), issue.get("actual", ""),
                issue.get("test_ref", ""), issue.get("status", "Open")]
        r = _row(ws, r, vals, alt=alt)

        # Color the severity cell based on level
        sev = issue.get("severity", "")
        if sev == "Critical":
            _sc(ws.cell(row=r - 1, column=2), FONT_FAIL, FILL_RED, A_C, BORDER)
        elif sev == "High":
            _sc(ws.cell(row=r - 1, column=2),
                Font(name="Calibri", bold=True, size=10, color="C65911"), FILL_GOLD, A_C, BORDER)
        elif sev == "Medium":
            _sc(ws.cell(row=r - 1, column=2), FONT_BOLD, FILL_LIGHT, A_C, BORDER)
        elif sev == "Low":
            _sc(ws.cell(row=r - 1, column=2), FONT_NORMAL, FILL_ALT, A_C, BORDER)
        else:
            _sc(ws.cell(row=r - 1, column=2), FONT_NORMAL, FILL_WHITE, A_C, BORDER)

    _widths(ws, {1: 5, 2: 12, 3: 16, 4: 18, 5: 50, 6: 40, 7: 40, 8: 14, 9: 14})


def _build_error_details(wb, results):
    ws = wb.create_sheet("Error Details")
    ws.sheet_properties.tabColor = "C00000"

    failed = [r for r in results if r.get("status") == "FAILED"]
    _title_banner(ws, "ERROR DETAILS", str(len(failed)) + " test(s) failed", cols=4)

    r = 4
    r = _headers(ws, r, ["#", "Test Name", "Module", "Error / Traceback"])

    for i, res in enumerate(failed, 1):
        alt = i % 2 == 0
        error = res.get("error", "No error details available")
        r = _row(ws, r, [i, res.get("test_name", ""), res.get("module", ""), error], alt=alt)
        _sc(ws.cell(row=r - 1, column=4), FONT_FAIL, FILL_ALT if alt else FILL_WHITE, A_L, BORDER)

    _widths(ws, {1: 5, 2: 45, 3: 16, 4: 80})


# -- Utilities ---------------------------------------------------

def _fmt_duration(seconds):
    if seconds is None:
        return "0s"
    s = int(seconds)
    if s < 60:
        return str(s) + "s"
    m, rem = divmod(s, 60)
    if m < 60:
        return str(m) + "m " + str(rem).zfill(2) + "s"
    h, m = divmod(m, 60)
    return str(h) + "h " + str(m).zfill(2) + "m " + str(rem).zfill(2) + "s"


def _truncate(text, max_len):
    text = (text or "").strip()
    if not text:
        return ""
    return text if len(text) <= max_len else text[:max_len - 3] + "..."


# -- In-memory store (used by conftest.py hooks) -----------------

class CSReportStore:
    """
    In-memory store for collecting test results during a pytest session.
    Used by conftest.py hooks - tests never touch this directly.
    """

    def __init__(self):
        self.results = []
        self.known_issues = []
        self._current = None
        self._logs = []
        self._start_time = None

    def record_issue(self, severity, module, category, description,
                     expected="", actual="", test_ref="", status="Open"):
        """
        Record a known issue / bug found during testing.

        Args:
            severity: Critical | High | Medium | Low | Info
            module: e.g. "UOM", "Currency"
            category: Backend | Frontend | Validation | UI | Data Integrity
            description: what the issue is
            expected: what should happen
            actual: what actually happens
            test_ref: e.g. "Test 12, Test 14"
            status: Open | Confirmed | Fixed | By Design
        """
        self.known_issues.append({
            "severity": severity,
            "module": module,
            "category": category,
            "description": description,
            "expected": expected,
            "actual": actual,
            "test_ref": test_ref,
            "status": status,
        })

    def start_test(self, test_name, nodeid):
        self._current = {
            "test_name": test_name,
            "module": _detect_module(nodeid),
            "steps": [],
            "status": "PASSED",
            "duration": 0,
            "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M"),
            "error": "",
        }
        self._logs = []
        self._start_time = datetime.now()

    def add_log_message(self, message):
        if self._current is not None:
            self._logs.append(message)

    def finish_test(self, status, error=""):
        if self._current is None:
            return
        self._current["steps"] = parse_steps_from_logs(self._logs)
        self._current["status"] = status
        if error:
            self._current["error"] = _sanitize(error)
        if self._start_time:
            self._current["duration"] = round(
                (datetime.now() - self._start_time).total_seconds(), 1
            )
        self.results.append(self._current)
        self._current = None
        self._logs = []
        self._start_time = None

    def has_results(self):
        return len(self.results) > 0