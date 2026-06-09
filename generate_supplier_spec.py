#!/usr/bin/env python3
"""Generate Supplier_Automation_Spec_Final.xlsx with 7 sheets."""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUTPUT = "/home/z/my-project/download/Supplier_Automation_Spec_Final.xlsx"

wb = openpyxl.Workbook()

# ── Shared styles ──────────────────────────────────────────────────
header_font = Font(bold=True, size=11, color="FFFFFF")
header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

pass_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")   # green
xfail_fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")   # orange
xpass_fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")   # gold

bug_confirmed_fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")  # red
bug_fixed_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")       # green

thin_border = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

prop_font = Font(bold=True, size=11)
value_font = Font(size=11)
body_align = Alignment(vertical="center", wrap_text=True)
center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)


def style_header(ws, num_cols, row=1):
    for c in range(1, num_cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border


def style_body(ws, num_cols, start_row=2):
    for row in ws.iter_rows(min_row=start_row, max_row=ws.max_row, min_col=1, max_col=num_cols):
        for cell in row:
            cell.border = thin_border
            cell.alignment = body_align


def auto_width(ws, min_w=10, max_w=50):
    for col_cells in ws.columns:
        col_letter = get_column_letter(col_cells[0].column)
        best = min_w
        for cell in col_cells:
            if cell.value:
                length = max(len(str(line)) for line in str(cell.value).split("\n"))
                if length > best:
                    best = length
        ws.column_dimensions[col_letter].width = min(best + 3, max_w)


# ════════════════════════════════════════════════════════════════════
# SHEET 1: Overview
# ════════════════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = "Overview"
ws1.append(["Property", "Value"])

overview_data = [
    ("Module", "Registration"),
    ("Screen", "Supplier"),
    ("URL", "/#/dynamic-screens/Supplier/Supplier"),
    ("Form Type", "Multi-Step Stepper Popup (3 steps)"),
    ("Framework", "Angular Material"),
    ("Login", "Rular@admin.com / Rular@12345678"),
    ("Facility", "RuralLife Producer Company (index 0)"),
    ("Browser", "Microsoft Edge (WebDriver)"),
    ("Python", "3.14.3"),
    ("pytest", "9.0.2"),
    ("Total Tests", "42"),
    ("Passed", "35"),
    ("XFAIL", "6"),
    ("XPASS", "1"),
    ("Failed", "0"),
    ("Automation Date", "2026-05-25"),
    ("Execution Time", "~30 min (estimated, first full run + re-runs)"),
    ("Report Generator", "CSReportStore + generate_cs_report"),
]

for prop, val in overview_data:
    ws1.append([prop, val])

style_header(ws1, 2)
for r in range(2, ws1.max_row + 1):
    ws1.cell(row=r, column=1).font = prop_font
    ws1.cell(row=r, column=2).font = value_font
    ws1.cell(row=r, column=1).border = thin_border
    ws1.cell(row=r, column=2).border = thin_border
    ws1.cell(row=r, column=1).alignment = body_align
    ws1.cell(row=r, column=2).alignment = body_align

ws1.column_dimensions["A"].width = 22
ws1.column_dimensions["B"].width = 55


# ════════════════════════════════════════════════════════════════════
# SHEET 2: Field Inventory
# ════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("Field Inventory")
ws2.append(["Step", "Field Name", "Type", "Required", "Max Length", "Options/Notes"])

field_data = [
    # Step 1 — Universal Fields
    ("1 — Universal Fields", "Party Reference", "mat-select", "No", "-", "Dynamic farmer list"),
    ("1 — Universal Fields", "Ownership Status", "mat-select", "Yes", "-", "Owned/Leased/Proprietorship/Partnership/LLP/PLC/Private Limited Company/Individual"),
    ("1 — Universal Fields", "Company Name", "text input", "Yes", "255", "BUG-001: No special char validation"),
    ("1 — Universal Fields", "PO Type", "mat-select", "Yes", "-", "Domestic/Import"),
    ("1 — Universal Fields", "Email", "text input", "No", "255", "BUG-002 (FIXED): Now validates email format"),
    ("1 — Universal Fields", "Phone Number", "number input", "Yes", "-", "BUG-003: Spinner controls (type=number)"),
    ("1 — Universal Fields", "Default Currency", "mat-select", "Yes", "-", "100+ currencies (INR default)"),
    ("1 — Universal Fields", "PAN Number", "text input", "Yes", "255", "BUG-004 (FIXED): Now validates PAN format"),
    ("1 — Universal Fields", "Is MSME Registered?", "toggle", "No", "-", "Default: No"),
    ("1 — Universal Fields", "Status", "toggle", "No", "-", "Default: Active"),
    # Step 1 — Additional Details
    ("1 — Additional Details", "Is GST Set Off", "toggle", "No", "-", "Default: Yes"),
    ("1 — Additional Details", "Is TDS Applicable", "toggle", "No", "-", "Default: No"),
    ("1 — Additional Details", "Contact Person Name", "text input", "No", "255", "Free text"),
    ("1 — Additional Details", "Office Number", "text input", "No", "255", "Free text"),
    ("1 — Additional Details", "Payment Terms", "mat-select", "No", "-", "21 Days/14 Days/7 Days/Wallet/RTGS/Advance/Immediate/60 Days/30 Days"),
    ("1 — Additional Details", "Delivery Terms", "mat-select", "No", "-", "Delivery/Spot"),
    ("1 — Additional Details", "Mode Of Delivery", "mat-select", "No", "-", "Air/Courier/Sea/Railway/Truck"),
    # Step 2 — Address Details
    ("2 — Address Details", "Address Type", "mat-select", "Yes", "-", "Shipping/Billing"),
    ("2 — Address Details", "Country", "mat-select", "Yes", "-", "Cascading (30 countries)"),
    ("2 — Address Details", "State", "mat-select", "Yes", "-", "Cascading from Country"),
    ("2 — Address Details", "District", "mat-select", "Yes", "-", "Cascading from State"),
    ("2 — Address Details", "Taluka", "mat-select", "Yes", "-", "Cascading from District"),
    ("2 — Address Details", "Village", "mat-select", "No", "-", "Cascading from Taluka"),
    ("2 — Address Details", "Address", "text input", "Yes", "255", "Free text"),
    ("2 — Address Details", "Pin Code", "text input", "No", "255", "6-digit numeric"),
    ("2 — Address Details", "GSTIN", "text input", "No", "255", "15-char GST format"),
    # Step 3 — Bank Details
    ("3 — Bank Details", "Bank Name", "text input", "No", "255", "Free text"),
    ("3 — Bank Details", "Branch", "text input", "No", "255", "Free text"),
    ("3 — Bank Details", "IFSC Code", "text input", "No", "255", "11-char format"),
    ("3 — Bank Details", "Account Type", "mat-select", "No", "-", "Current/Saving"),
    ("3 — Bank Details", "Account Holder Name", "text input", "No", "255", "Free text"),
    ("3 — Bank Details", "Account Number", "text input", "No", "255", "Numeric"),
    ("3 — Bank Details", "Bank Proof", "mat-select", "Yes", "-", "Cancelled Cheque/Passbook"),
    ("3 — Bank Details", "Attachment", "file upload", "No", "-", ".png/.jpg/.pdf"),
]

for row_data in field_data:
    ws2.append(row_data)

style_header(ws2, 6)
style_body(ws2, 6)
ws2.column_dimensions["A"].width = 25
ws2.column_dimensions["B"].width = 25
ws2.column_dimensions["C"].width = 15
ws2.column_dimensions["D"].width = 12
ws2.column_dimensions["E"].width = 14
ws2.column_dimensions["F"].width = 55


# ════════════════════════════════════════════════════════════════════
# SHEET 3: Test Plan (with color coding + auto-filter)
# ════════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("Test Plan")
ws3.append(["Test ID", "Test Name", "Class", "Priority", "Description", "Expected Result", "Status", "Notes"])

test_plan = [
    # Phase 1 — Create Form Validations (18 tests)
    ("SP-C01", "test_SP_C01_empty_submit", "TestCreateFormValidations", "Critical", "Submit with all required fields empty", "SweetAlert2/mat-errors or form stays open", "PASSED", "Validation warning shown"),
    ("SP-C02", "test_SP_C02_valid_create", "TestCreateFormValidations", "Critical", "Create with valid data across all 3 steps", "Success SweetAlert2 + table entry", "PASSED", "Re-run passed (env issue first run)"),
    ("SP-C03", "test_SP_C03_company_name_spaces", "TestCreateFormValidations", "High", "Spaces-only Company Name", "Rejected or validation error", "PASSED", "Caught as empty"),
    ("SP-C04", "test_SP_C04_company_name_special_chars", "TestCreateFormValidations", "High", "Special chars in Company Name", "Rejected (BUG-001)", "XFAIL", "BUG-001: accepted"),
    ("SP-C05", "test_SP_C05_company_name_sql_injection", "TestCreateFormValidations", "Critical", "SQL injection in Company Name", "Rejected (BUG-001)", "XFAIL", "BUG-001: accepted"),
    ("SP-C06", "test_SP_C06_company_name_xss", "TestCreateFormValidations", "Critical", "XSS payload in Company Name", "Rejected (BUG-001)", "XFAIL", "BUG-001: accepted"),
    ("SP-C07", "test_SP_C07_company_name_255_chars", "TestCreateFormValidations", "Medium", "255-char Company Name", "Accepted (maxlength boundary)", "PASSED", "maxlength working"),
    ("SP-C08", "test_SP_C08_company_name_256_chars", "TestCreateFormValidations", "Medium", "256-char Company Name", "Truncated to 255", "PASSED", "maxlength enforced"),
    ("SP-C09", "test_SP_C09_invalid_email", "TestCreateFormValidations", "High", "Invalid email format", "Rejected with error", "PASSED", "BUG-002 FIXED"),
    ("SP-C10", "test_SP_C10_invalid_pan", "TestCreateFormValidations", "High", "Invalid PAN format", "Rejected with error", "PASSED", "BUG-004 FIXED"),
    ("SP-C11", "test_SP_C11_phone_alpha_chars", "TestCreateFormValidations", "Medium", "Alpha chars in Phone Number", "Rejected (type=number)", "PASSED", "type=number rejects alpha"),
    ("SP-C12", "test_SP_C12_ownership_status_dropdown", "TestCreateFormValidations", "Medium", "Ownership Status options", "Shows 8 options", "PASSED", "Verified all options"),
    ("SP-C13", "test_SP_C13_po_type_dropdown", "TestCreateFormValidations", "Medium", "PO Type options", "Domestic/Import shown", "PASSED", "Re-run passed (env issue first run)"),
    ("SP-C14", "test_SP_C14_currency_dropdown", "TestCreateFormValidations", "Medium", "Currency options", "100+ currencies", "PASSED", "INR confirmed"),
    ("SP-C15", "test_SP_C15_payment_terms_dropdown", "TestCreateFormValidations", "Low", "Payment Terms options", "Options shown", "PASSED", "Scrolled to Additional Details"),
    ("SP-C16", "test_SP_C16_delivery_terms_dropdown", "TestCreateFormValidations", "Low", "Delivery Terms options", "Options shown", "PASSED", "Re-run passed (env issue first run)"),
    ("SP-C17", "test_SP_C17_mode_of_delivery_dropdown", "TestCreateFormValidations", "Low", "Mode Of Delivery options", "Air/Courier/Sea/Railway/Truck", "PASSED", "Verified"),
    ("SP-C18", "test_SP_C18_stepper_navigation", "TestCreateFormValidations", "Critical", "Next/Back navigation", "Steps advance and return correctly", "PASSED", "Step 0→1→0 verified"),
    # Phase 2 — Duplicate Validations (3 tests)
    ("SP-D01", "test_SP_D01_duplicate_company_name", "TestDuplicateValidations", "High", "Duplicate Company Name", "Check behavior", "PASSED", "Duplicate allowed (no uniqueness check)"),
    ("SP-D02", "test_SP_D02_duplicate_email", "TestDuplicateValidations", "High", "Duplicate Email", "Check behavior", "PASSED", "Duplicate allowed"),
    ("SP-D03", "test_SP_D03_duplicate_phone", "TestDuplicateValidations", "High", "Duplicate Phone Number", "Check behavior", "PASSED", "Duplicate allowed"),
    # Phase 3 — Edit Form Validations (4 tests)
    ("SP-E01", "test_SP_E01_edit_no_update_button", "TestEditFormValidations", "High", "Edit popup has Update button", "Update button visible", "PASSED", "BUG-005 FIXED"),
    ("SP-E02", "test_SP_E02_edit_prepopulated", "TestEditFormValidations", "High", "Edit fields pre-populated", "Company Name pre-filled", "PASSED", "Verified pre-fill"),
    ("SP-E03", "test_SP_E03_edit_company_name_special_chars", "TestEditFormValidations", "High", "Edit to special chars", "Rejected (BUG-001)", "XPASS", "BUG-001 STILL ACTIVE: special chars accepted in Edit"),
    ("SP-E04", "test_SP_E04_edit_invalid_email", "TestEditFormValidations", "Medium", "Edit to invalid email", "Rejected", "PASSED", "Email validation works in Edit"),
    # Phase 4 — Search & Filter (5 tests)
    ("SP-S01", "test_SP_S01_search_exact", "TestSearchFilter", "High", "Exact Company Name search", "Supplier found", "PASSED", "Match confirmed"),
    ("SP-S02", "test_SP_S02_search_partial", "TestSearchFilter", "High", "Partial Company Name search", "Suppliers found", "PASSED", "First 10 chars matched"),
    ("SP-S03", "test_SP_S03_search_case_insensitive", "TestSearchFilter", "Medium", "Case-insensitive search", "Results found", "PASSED", "Lowercase search works"),
    ("SP-S04", "test_SP_S04_search_no_results", "TestSearchFilter", "Medium", "Non-existent search", "No results", "PASSED", "No false positives"),
    ("SP-S05", "test_SP_S05_search_special_chars", "TestSearchFilter", "Low", "Special char search", "No crash", "PASSED", "Stable with special chars"),
    # Phase 5 — Popup & UI Behaviors (7 tests)
    ("SP-P01", "test_SP_P01_add_form_opens", "TestPopupUIBehaviors", "Critical", "ADD opens stepper popup", "Popup visible + stepper present", "PASSED", "3-step stepper confirmed"),
    ("SP-P02", "test_SP_P02_view_popup_readonly", "TestPopupUIBehaviors", "High", "View popup read-only", "All fields disabled", "PASSED", "Read-only verified"),
    ("SP-P03", "test_SP_P03_cancel_closes_popup", "TestPopupUIBehaviors", "High", "Cancel closes without creating", "Row count unchanged", "PASSED", "No data leakage"),
    ("SP-P04", "test_SP_P04_close_button", "TestPopupUIBehaviors", "High", "X button closes popup", "Row count unchanged", "PASSED", "Clean close"),
    ("SP-P05", "test_SP_P05_sweetalert_success", "TestPopupUIBehaviors", "Critical", "SweetAlert2 on valid create", "Success toast shown", "PASSED", "SweetAlert2 confirmed"),
    ("SP-P06", "test_SP_P06_phone_spinner_controls", "TestPopupUIBehaviors", "Low", "Phone spinner check", "No spinner visible", "XFAIL", "BUG-003: spinner still present"),
    ("SP-P07", "test_SP_P07_toggle_defaults", "TestPopupUIBehaviors", "Medium", "Toggle default values", "MSME=No/Status=Active/GST=Yes/TDS=No", "PASSED", "All defaults verified"),
    # Phase 6 — Bug-Specific (5 tests)
    ("SP-B01", "test_SP_B01_special_chars_accepted", "TestBugSpecific", "Critical", "BUG-001 special chars create", "Rejected", "XFAIL", "BUG-001 CONFIRMED: accepted"),
    ("SP-B02", "test_SP_B02_no_email_validation", "TestBugSpecific", "High", "BUG-002 invalid email create", "Accepted", "PASSED", "BUG-002 FIXED: now rejected"),
    ("SP-B03", "test_SP_B03_phone_spinner", "TestBugSpecific", "Low", "BUG-003 phone spinner", "No spinner", "XFAIL", "BUG-003 CONFIRMED: spinner present"),
    ("SP-B04", "test_SP_B04_no_pan_validation", "TestBugSpecific", "High", "BUG-004 invalid PAN create", "Accepted", "PASSED", "BUG-004 FIXED: now rejected"),
    ("SP-B05", "test_SP_B05_edit_no_update", "TestBugSpecific", "High", "BUG-005 edit has Update", "Update visible", "PASSED", "BUG-005 FIXED + Re-run passed"),
]

for row_data in test_plan:
    ws3.append(row_data)

style_header(ws3, 8)
style_body(ws3, 8)

# Color-code the Status column (column 7)
status_col = 7
for r in range(2, ws3.max_row + 1):
    status = ws3.cell(row=r, column=status_col).value
    if status == "PASSED":
        ws3.cell(row=r, column=status_col).fill = pass_fill
    elif status == "XFAIL":
        ws3.cell(row=r, column=status_col).fill = xfail_fill
    elif status == "XPASS":
        ws3.cell(row=r, column=status_col).fill = xpass_fill
    # Center-align Test ID and Status columns
    ws3.cell(row=r, column=1).alignment = center_align
    ws3.cell(row=r, column=status_col).alignment = center_align
    ws3.cell(row=r, column=4).alignment = center_align  # Priority

# Auto-filter
ws3.auto_filter.ref = f"A1:H{ws3.max_row}"

auto_width(ws3)


# ════════════════════════════════════════════════════════════════════
# SHEET 4: Known Bugs
# ════════════════════════════════════════════════════════════════════
ws4 = wb.create_sheet("Known Bugs")
ws4.append(["Bug ID", "Severity", "Category", "Description", "Expected", "Actual", "Status", "Test Ref"])

bug_data = [
    ("BUG-001", "High", "Validation", "Company Name accepts special characters (@#$%^&*)", "Should restrict and show error", "CONFIRMED: 'Test@#Traders&Co!' saved", "Confirmed", "SP-C04/C05/C06/E03/B01"),
    ("BUG-002", "Medium", "Validation", "No email format validation", "Should validate email format", "FIXED: Now shows error", "Fixed", "SP-C09/B02"),
    ("BUG-003", "Low", "UI Bug", "Phone Number has spinner controls (type=number)", "Should be type=tel", "Spinner still visible", "Confirmed", "SP-P06/B03"),
    ("BUG-004", "Medium", "Validation", "No PAN format validation", "Should validate PAN format", "FIXED: Now shows error", "Fixed", "SP-C10/B04"),
    ("BUG-005", "High", "Functionality", "No Update button in Edit mode", "Update button should exist", "FIXED: Now visible", "Fixed", "SP-E01/B05"),
]

for row_data in bug_data:
    ws4.append(row_data)

style_header(ws4, 8)
style_body(ws4, 8)

# Color-code Status column (column 7)
status_col = 7
for r in range(2, ws4.max_row + 1):
    status = ws4.cell(row=r, column=status_col).value
    if status == "Confirmed":
        ws4.cell(row=r, column=status_col).fill = bug_confirmed_fill
    elif status == "Fixed":
        ws4.cell(row=r, column=status_col).fill = bug_fixed_fill
    ws4.cell(row=r, column=status_col).alignment = center_align
    ws4.cell(row=r, column=1).alignment = center_align

auto_width(ws4)


# ════════════════════════════════════════════════════════════════════
# SHEET 5: Test Summary
# ════════════════════════════════════════════════════════════════════
ws5 = wb.create_sheet("Test Summary")
ws5.append(["Metric", "Value"])

summary_data = [
    ("Total Tests", "42"),
    ("PASSED", "35 (83.3%)"),
    ("XFAIL", "6 (14.3%) — Expected failures (confirmed bugs)"),
    ("XPASS", "1 (2.4%) — Unexpected pass (bug still present: SP-E03)"),
    ("FAILED", "0 (0%)"),
    ("", ""),
    ("Note", "First run had 3 FAILED + 1 ERROR due to environment/session issues. All passed on re-run."),
    ("Environment Failures", "SP-C02 (no submit response — timeout), SP-C13/SP-C16 (dropdown not loaded — timing), SP-B05 (InvalidSessionId — browser session expired after long run)"),
]

for row_data in summary_data:
    ws5.append(row_data)

style_header(ws5, 2)
for r in range(2, ws5.max_row + 1):
    ws5.cell(row=r, column=1).font = prop_font
    ws5.cell(row=r, column=2).font = value_font
    ws5.cell(row=r, column=1).border = thin_border
    ws5.cell(row=r, column=2).border = thin_border
    ws5.cell(row=r, column=1).alignment = body_align
    ws5.cell(row=r, column=2).alignment = body_align

ws5.column_dimensions["A"].width = 22
ws5.column_dimensions["B"].width = 90


# ════════════════════════════════════════════════════════════════════
# SHEET 6: Action Flows
# ════════════════════════════════════════════════════════════════════
ws6 = wb.create_sheet("Action Flows")
ws6.append(["Flow Name", "Steps", "Page Methods"])

flow_data = [
    ("Create Supplier (Happy Path)",
     "Open ADD → Fill Step 1 Universal → Scroll to Additional Details → Fill Step 1 Additional → Click Next → Fill Step 2 Address (cascading) → Click Next → Fill Step 3 Bank → Click Submit → Handle SweetAlert2",
     "open_add_form, fill_step1_universal, scroll_to_additional_details, fill_step1_additional, click_stepper_next, fill_step2_address, fill_step3_bank, click_submit, handle_success_alert"),
    ("View Supplier",
     "Click row menu → Click View → Verify read-only → Close",
     "click_view_first_row, verify_view_popup_read_only, close_popup"),
    ("Edit Supplier",
     "Click row menu → Click Edit → Verify pre-populated → Modify fields → Click Update → Handle SweetAlert2",
     "click_edit_first_row, get_form_field_values, has_update_button, click_update, handle_success_alert"),
    ("Search Supplier",
     "Click search toggle → Type search text → Verify results → Clear search",
     "search_supplier, click_refresh"),
    ("Cancel Form",
     "Open ADD → Fill partial data → Click Cancel → Verify no data saved",
     "open_add_form, fill_step1_universal, cancel, get_table_row_count"),
]

for row_data in flow_data:
    ws6.append(row_data)

style_header(ws6, 3)
style_body(ws6, 3)
ws6.column_dimensions["A"].width = 30
ws6.column_dimensions["B"].width = 80
ws6.column_dimensions["C"].width = 65


# ════════════════════════════════════════════════════════════════════
# SHEET 7: Selector Map
# ════════════════════════════════════════════════════════════════════
ws7 = wb.create_sheet("Selector Map")
ws7.append(["Element", "Locator Strategy", "Locator Value", "Notes"])

selector_data = [
    # Toolbar
    ("ADD button", "CSS", "button.erp-add-btn", "Mini-fab with add icon"),
    ("Search toggle", "CSS", "button.search-btn", "button[mattooltip='Search']"),
    ("Search input", "CSS", "#erpSearchInput", "input.erp-search-input"),
    ("Refresh button", "CSS", "button[mattooltip='Refresh']", ""),
    # Table
    ("Table", "CSS", "table#excel-table", "table.mat-mdc-table"),
    ("Table rows", "CSS", "table#excel-table tbody tr", ""),
    ("Name cells", "CSS", "td.cdk-column-name", ""),
    ("Phone cells", "CSS", "td.cdk-column-mobile_no", ""),
    ("Status cells", "CSS", "td.cdk-column-status", ""),
    # Row Actions
    ("Row menu trigger", "CSS", "button.mat-mdc-menu-trigger.erp-row-trigger", ""),
    ("View menu item", "XPATH", "//button[.//i[contains(text(),'visibility')]]", ""),
    ("Edit menu item", "XPATH", "//button[.//i[contains(text(),'edit')]]", ""),
    # Form Popup
    ("Form popup", "CSS", ".big-model, .edit_pop_up, mat-dialog-container", ""),
    ("Stepper", "CSS", "mat-stepper, mat-horizontal-stepper", ""),
    ("Step headers", "CSS", "mat-step-header", ""),
    # Step 1 Inputs
    ("Company Name", "CSS", "input[name='Company Name']", ""),
    ("Email", "CSS", "input[name='Email']", ""),
    ("Phone Number", "CSS", "input[name='Phone Number']", ""),
    ("PAN Number", "CSS", "input[name='PAN Number']", ""),
    ("Contact Person Name", "CSS", "input[name='Contact Person Name']", ""),
    ("Office Number", "CSS", "input[name='Office Number']", ""),
    # Step 1 Selects (XPath — label-based)
    ("Party Reference", "XPATH", "//mat-label[contains(.,'Party Reference')]/ancestor::mat-form-field//mat-select", ""),
    ("Ownership Status", "XPATH", "//mat-label[contains(.,'Ownership Status')]/ancestor::mat-form-field//mat-select", ""),
    ("PO Type", "XPATH", "//mat-label[contains(.,'PO Type')]/ancestor::mat-form-field//mat-select", ""),
    ("Default Currency", "XPATH", "//mat-label[contains(.,'Default Currency')]/ancestor::mat-form-field//mat-select", ""),
    ("Payment Terms", "XPATH", "//mat-label[contains(.,'Payment Terms')]/ancestor::mat-form-field//mat-select", ""),
    ("Delivery Terms", "XPATH", "//mat-label[contains(.,'Delivery Terms')]/ancestor::mat-form-field//mat-select", ""),
    ("Mode Of Delivery", "XPATH", "//mat-label[contains(.,'Mode Of Delivery')]/ancestor::mat-form-field//mat-select", ""),
    # Step 2 Selects
    ("Address Type", "XPATH", "//mat-label[contains(.,'Address Type')]/ancestor::mat-form-field//mat-select", ""),
    ("Country", "XPATH", "//mat-label[contains(.,'Country')]/ancestor::mat-form-field//mat-select", ""),
    ("State", "XPATH", "//mat-label[contains(.,'State')]/ancestor::mat-form-field//mat-select", ""),
    ("District", "XPATH", "//mat-label[contains(.,'District')]/ancestor::mat-form-field//mat-select", ""),
    ("Taluka", "XPATH", "//mat-label[contains(.,'Taluka')]/ancestor::mat-form-field//mat-select", ""),
    ("Village", "XPATH", "//mat-label[contains(.,'Village')]/ancestor::mat-form-field//mat-select", ""),
    ("Address", "CSS", "input[name='Address']", ""),
    ("Pin Code", "CSS", "input[name='Pin Code']", ""),
    ("GSTIN", "CSS", "input[name='GSTIN']", ""),
    # Step 3
    ("Bank Name", "CSS", "input[name='Bank Name']", ""),
    ("Branch", "CSS", "input[name='Branch']", ""),
    ("IFSC Code", "CSS", "input[name='IFSC Code']", ""),
    ("Account Holder Name", "CSS", "input[name='Account Holder Name']", ""),
    ("Account Number", "CSS", "input[name='Account Number']", ""),
    ("Account Type", "XPATH", "//mat-label[contains(.,'Account Type')]/ancestor::mat-form-field//mat-select", ""),
    ("Bank Proof", "XPATH", "//mat-label[contains(.,'Bank Proof')]/ancestor::mat-form-field//mat-select", ""),
    # Navigation Buttons
    ("Next", "CSS", "button.mat-stepper-next", ""),
    ("Back", "CSS", "button.mat-stepper-previous", ""),
    ("Submit", "XPATH", "//div[contains(@class,'popup-footer')]//button[contains(.,'Submit')]", ""),
    ("Update", "XPATH", "//div[contains(@class,'popup-footer')]//button[contains(.,'Update')]", ""),
    ("Cancel", "XPATH", "//div[contains(@class,'popup-footer')]//button[contains(.,'Cancel')]", ""),
    # SweetAlert2
    ("Title", "CSS", "#swal2-title", ""),
    ("HTML container", "CSS", ".swal2-html-container", ""),
    ("Confirm button", "CSS", ".swal2-confirm", ""),
    ("Cancel button", "CSS", ".swal2-cancel", ""),
]

for row_data in selector_data:
    ws7.append(row_data)

style_header(ws7, 4)
style_body(ws7, 4)
ws7.column_dimensions["A"].width = 25
ws7.column_dimensions["B"].width = 18
ws7.column_dimensions["C"].width = 80
ws7.column_dimensions["D"].width = 30


# ── Save ────────────────────────────────────────────────────────────
wb.save(OUTPUT)
print(f"✅ File saved: {OUTPUT}")
print(f"   Sheets: {wb.sheetnames}")
print(f"   Total rows in Test Plan: {ws3.max_row - 1}")
